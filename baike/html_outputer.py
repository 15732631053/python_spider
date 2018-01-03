#!/usr/bin/env python2
# -*- coding: UTF-8 -*-
from openpyxl import load_workbook
class HtmlOutputer(object):

    def __init__(self):
        self.datas = []


    def collect_data(self, data):
        if data is None:
            return
        self.datas.append(data)

    def output_html(self):
        fout = open('output.html', 'w')
        fout.write("<html>")
        fout.write("<body>")
        fout.write("<table>")

        for data in self.datas:
            fout.write("<tr>")
            fout.write("<td ><a href=%s>链接</a></td>" % data['url'].encode('utf-8'))
            fout.write("<td>%s</td>" % data['title'].encode('utf-8'))
            fout.write("<td>%s</td>" % data['summary'].encode('utf-8'))
            fout.write("</tr>")
        fout.write("</table>")
        fout.write("</body>")
        fout.write("</html>")
        fout.close()
    def output_excel(self):
        #操作excel表
        wb = load_workbook('balances.xlsx')
        ws = wb.active
        for i,data in enumerate(self.datas):
            ws['A%d' %(i+2)]=data['url'];
            ws['B%d' %(i+2)]=data['title'].encode('utf-8');
            ws['C%d' %(i+2)]=data['summary'].encode('utf-8');

        wb.save('balances.xlsx')

